import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


SAMPLE_DIR = Path(__file__).parent / "sample_data"
SAMPLE_DIR.mkdir(exist_ok=True)


def _set_cell(ws, row, col, value, merge_range=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if merge_range:
        ws.merge_cells(merge_range)


def generate_store_a():
    wb = Workbook()
    ws = wb.active
    ws.title = "1月排班"

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="旗舰店2025年1月排班表")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    headers = ["员工号", "姓名", "日期", "班次", "上班时间", "下班时间"]
    for i, h in enumerate(headers, 1):
        _set_cell(ws, 2, i, h)

    data = [
        ("E001", "张三", datetime.date(2025, 1, 6), "A", "09:00", "17:00"),
        ("E001", "张三", datetime.date(2025, 1, 7), "A", "09:00", "17:00"),
        ("E001", "张三", datetime.date(2025, 1, 8), "B", "13:00", "21:00"),
        ("E001", "张三", datetime.date(2025, 1, 9), "A", "09:00", "17:00"),
        ("E001", "张三", datetime.date(2025, 1, 10), "A", "09:00", "17:00"),
        ("E001", "张三", datetime.date(2025, 1, 11), "C", "21:00", "05:00"),
        ("E001", "张三", datetime.date(2025, 1, 12), "A", "09:00", "17:00"),
        ("E002", "李四", datetime.date(2025, 1, 6), "B", "13:00", "21:00"),
        ("E002", "李四", datetime.date(2025, 1, 7), "B", "13:00", "21:00"),
        ("E002", "李四", datetime.date(2025, 1, 8), "OFF", None, None),
        ("E002", "李四", datetime.date(2025, 1, 9), "B", "13:00", "21:00"),
        ("E002", "李四", datetime.date(2025, 1, 10), "B", "13:00", "21:00"),
        ("E002", "李四", datetime.date(2025, 1, 11), "B", "13:00", "21:00"),
        ("E002", "李四", datetime.date(2025, 1, 12), "B", "13:00", "21:00"),
        ("E003", "王五", datetime.date(2025, 1, 6), "A", "09:00", "17:00"),
        ("E003", "王五", datetime.date(2025, 1, 7), "A", "09:00", "17:00"),
        ("E003", "王五", datetime.date(2025, 1, 8), "A", "09:00", "17:00"),
        ("E003", "王五", datetime.date(2025, 1, 9), "A", "09:00", "17:00"),
        ("E003", "王五", datetime.date(2025, 1, 10), "A", "09:00", "17:00"),
        ("E003", "王五", datetime.date(2025, 1, 11), "C", "21:00", "05:00"),
        ("E003", "王五", datetime.date(2025, 1, 12), "A", "06:00", "14:00"),
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            _set_cell(ws, r, c, val)

    wb.save(SAMPLE_DIR / "旗舰店_202501.xlsx")


def generate_store_b():
    wb = Workbook()
    ws = wb.active
    ws.title = "排班"

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="分店A 一月份排班")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    headers = ["工号", "名字", "工作日", "班别"]
    for i, h in enumerate(headers, 1):
        _set_cell(ws, 2, i, h)

    data = [
        ("E004", "赵六", "2025/1/6", "A"),
        ("E004", "赵六", "2025/1/7", "A"),
        ("E004", "赵六", "2025/1/8", "C"),
        ("E004", "赵六", "2025/1/9", "休"),
        ("E004", "赵六", "2025/1/10", "A"),
        ("E004", "赵六", "2025/1/11", "A"),
        ("E004", "赵六", "2025/1/12", "A"),
        ("E005", "孙七", "2025/1/6", "早"),
        ("E005", "孙七", "2025/1/7", "早"),
        ("E005", "孙七", "2025/1/8", "早"),
        ("E005", "孙七", "2025/1/9", "早"),
        ("E005", "孙七", "2025/1/10", "晚"),
        ("E005", "孙七", "2025/1/11", "早"),
        ("E005", "孙七", "2025/1/12", "早"),
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            _set_cell(ws, r, c, val)

    wb.save(SAMPLE_DIR / "分店A_202501.xlsx")


def generate_store_c():
    wb = Workbook()
    ws = wb.active
    ws.title = "1月"

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="分店B·2025年1月排班记录")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    headers = ["EmpID", "Name", "Date", "签到", "签退"]
    for i, h in enumerate(headers, 1):
        _set_cell(ws, 2, i, h)

    data = [
        ("E006", "周八", datetime.datetime(2025, 1, 6), "09:00", "17:00"),
        ("E006", "周八", datetime.datetime(2025, 1, 7), "13:00", "21:00"),
        ("E006", "周八", datetime.datetime(2025, 1, 8), "09:00", "17:00"),
        ("E006", "周八", datetime.datetime(2025, 1, 9), "09:00", "17:00"),
        ("E006", "周八", datetime.datetime(2025, 1, 10), "21:00", "05:00"),
        ("E006", "周八", datetime.datetime(2025, 1, 11), "09:00", "17:00"),
        ("E006", "周八", datetime.datetime(2025, 1, 12), "09:00", "17:00"),
        ("E001", "zhangsan", datetime.datetime(2025, 1, 8), "09:00", "17:00"),
        ("E001", "zhangsan", datetime.datetime(2025, 1, 9), "09:00", "17:00"),
        ("E001", "zhangsan", datetime.datetime(2025, 1, 10), "13:00", "21:00"),
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            _set_cell(ws, r, c, val)

    wb.save(SAMPLE_DIR / "分店B_202501.xlsx")


def generate_store_d_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "排班表"

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="分店C 2025年1月排班（含合并单元格）")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    headers = ["员工编号", "姓名", "1月13日", "1月14日", "1月15日", "1月16日", "1月17日"]
    for i, h in enumerate(headers, 1):
        _set_cell(ws, 2, i, h)

    ws.merge_cells("A3:A4")
    _set_cell(ws, 3, 1, "E007", "A3:A4")
    ws.merge_cells("B3:B4")
    _set_cell(ws, 3, 2, "吴九", "B3:B4")

    row3 = ["早", "中", "晚", "休", "早"]
    row4 = ["中", "早", "休", "早", "中"]
    for i, v in enumerate(row3):
        _set_cell(ws, 3, 3 + i, v)
    for i, v in enumerate(row4):
        _set_cell(ws, 4, 3 + i, v)

    wb.save(SAMPLE_DIR / "分店C_202501.xlsx")


if __name__ == "__main__":
    generate_store_a()
    generate_store_b()
    generate_store_c()
    generate_store_d_merged_cells()
    print(f"Sample files generated in {SAMPLE_DIR}")
