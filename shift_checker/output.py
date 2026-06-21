from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from shift_checker.compliance import Violation
from shift_checker.cost import (
    EmployeeCostSummary,
    ShiftCost,
    StoreCostSummary,
)
from shift_checker.parser import ShiftRecord


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CRITICAL_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row: int, max_col: int):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, max_col: int, min_width: int = 10, max_width: int = 30):
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def write_merged_table(
    records: list[ShiftRecord],
    output_path: str | Path,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "合并总表"

    headers = [
        "员工号", "姓名", "门店", "日期", "班次",
        "上班时间", "下班时间", "工时", "跨天", "来源文件",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    sorted_records = sorted(records, key=lambda r: (r.employee_id, r.date))
    for idx, r in enumerate(sorted_records, 2):
        ws.cell(row=idx, column=1, value=r.employee_id)
        ws.cell(row=idx, column=2, value=r.employee_name)
        ws.cell(row=idx, column=3, value=r.store)
        ws.cell(row=idx, column=4, value=r.date.strftime("%Y-%m-%d"))
        ws.cell(row=idx, column=5, value=r.shift_code)
        ws.cell(row=idx, column=6, value=r.start_time.strftime("%H:%M") if r.start_time else "")
        ws.cell(row=idx, column=7, value=r.end_time.strftime("%H:%M") if r.end_time else "")
        ws.cell(row=idx, column=8, value=r.hours)
        ws.cell(row=idx, column=9, value="是" if r.overnight else "否")
        ws.cell(row=idx, column=10, value=Path(r.source_file).name)
        for c in range(1, len(headers) + 1):
            ws.cell(row=idx, column=c).border = THIN_BORDER

    _auto_width(ws, len(headers))
    wb.save(output_path)
    wb.close()


def write_violations(
    violations: list[Violation],
    output_path: str | Path,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "违规清单"

    headers = ["员工号", "姓名", "违规类型", "严重程度", "详情", "涉及日期"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    for idx, v in enumerate(violations, 2):
        ws.cell(row=idx, column=1, value=v.employee_id)
        ws.cell(row=idx, column=2, value=v.employee_name)
        ws.cell(row=idx, column=3, value=v.rule)
        ws.cell(row=idx, column=4, value=v.severity)
        ws.cell(row=idx, column=5, value=v.details)
        ws.cell(row=idx, column=6, value=", ".join(d.strftime("%Y-%m-%d") for d in v.dates))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=c)
            cell.border = THIN_BORDER
        sev_cell = ws.cell(row=idx, column=4)
        if v.severity == "critical":
            sev_cell.fill = CRITICAL_FILL
        elif v.severity == "warning":
            sev_cell.fill = WARNING_FILL

    _auto_width(ws, len(headers))
    wb.save(output_path)
    wb.close()


def write_cost_summary(
    emp_summaries: list[EmployeeCostSummary],
    store_summaries: list[StoreCostSummary],
    output_path: str | Path,
):
    wb = Workbook()

    ws_emp = wb.active
    ws_emp.title = "员工成本"
    emp_headers = [
        "员工号", "姓名", "门店", "岗位",
        "总工时", "常规工时", "加班工时",
        "常规成本", "加班成本", "总成本",
    ]
    for i, h in enumerate(emp_headers, 1):
        ws_emp.cell(row=1, column=i, value=h)
    _style_header(ws_emp, 1, len(emp_headers))

    for idx, s in enumerate(emp_summaries, 2):
        ws_emp.cell(row=idx, column=1, value=s.employee_id)
        ws_emp.cell(row=idx, column=2, value=s.employee_name)
        ws_emp.cell(row=idx, column=3, value=s.store)
        ws_emp.cell(row=idx, column=4, value=s.position)
        ws_emp.cell(row=idx, column=5, value=round(s.total_hours, 1))
        ws_emp.cell(row=idx, column=6, value=round(s.regular_hours, 1))
        ws_emp.cell(row=idx, column=7, value=round(s.overtime_hours, 1))
        ws_emp.cell(row=idx, column=8, value=round(s.regular_cost, 2))
        ws_emp.cell(row=idx, column=9, value=round(s.overtime_cost, 2))
        ws_emp.cell(row=idx, column=10, value=round(s.total_cost, 2))
        for c in range(1, len(emp_headers) + 1):
            ws_emp.cell(row=idx, column=c).border = THIN_BORDER

    _auto_width(ws_emp, len(emp_headers))

    ws_store = wb.create_sheet("门店成本")
    store_headers = ["门店", "总工时", "常规成本", "加班成本", "总成本", "人数"]
    for i, h in enumerate(store_headers, 1):
        ws_store.cell(row=1, column=i, value=h)
    _style_header(ws_store, 1, len(store_headers))

    for idx, s in enumerate(store_summaries, 2):
        ws_store.cell(row=idx, column=1, value=s.store)
        ws_store.cell(row=idx, column=2, value=round(s.total_hours, 1))
        ws_store.cell(row=idx, column=3, value=round(s.regular_cost, 2))
        ws_store.cell(row=idx, column=4, value=round(s.overtime_cost, 2))
        ws_store.cell(row=idx, column=5, value=round(s.total_cost, 2))
        ws_store.cell(row=idx, column=6, value=s.employee_count)
        for c in range(1, len(store_headers) + 1):
            ws_store.cell(row=idx, column=c).border = THIN_BORDER

    _auto_width(ws_store, len(store_headers))
    wb.save(output_path)
    wb.close()


def generate_markdown_report(
    records: list[ShiftRecord],
    violations: list[Violation],
    emp_summaries: list[EmployeeCostSummary],
    store_summaries: list[StoreCostSummary],
    output_path: str | Path,
):
    lines: list[str] = []
    lines.append("# 排班合规与成本分析报告\n")

    lines.append("## 概览\n")
    lines.append(f"- 排班记录总数: **{len(records)}**")
    emp_ids = set(r.employee_id for r in records)
    stores = set(r.store for r in records)
    lines.append(f"- 涉及员工: **{len(emp_ids)}** 人")
    lines.append(f"- 涉及门店: **{len(stores)}** 家 ({', '.join(sorted(stores))})")
    lines.append(f"- 违规总数: **{len(violations)}** 条")
    critical = sum(1 for v in violations if v.severity == "critical")
    warning = sum(1 for v in violations if v.severity == "warning")
    lines.append(f"  - 严重: {critical} 条")
    lines.append(f"  - 警告: {warning} 条")
    total_cost = sum(s.total_cost for s in emp_summaries)
    total_ot = sum(s.overtime_cost for s in emp_summaries)
    lines.append(f"- 总人力成本: **¥{total_cost:,.2f}**")
    lines.append(f"- 其中加班成本: **¥{total_ot:,.2f}**")
    lines.append("")

    if violations:
        lines.append("## 违规详情\n")
        by_rule: dict[str, list[Violation]] = {}
        for v in violations:
            by_rule.setdefault(v.rule, []).append(v)
        for rule, vs in sorted(by_rule.items()):
            lines.append(f"### {rule} ({len(vs)} 条)\n")
            for v in vs:
                dates_str = ", ".join(d.strftime("%m/%d") for d in v.dates)
                lines.append(f"- **{v.employee_name}** ({v.employee_id}): {v.details} [{dates_str}]")
            lines.append("")

    lines.append("## 门店成本汇总\n")
    lines.append("| 门店 | 总工时 | 常规成本 | 加班成本 | 总成本 | 人数 |")
    lines.append("|------|--------|----------|----------|--------|------|")
    for s in store_summaries:
        lines.append(
            f"| {s.store} | {s.total_hours:.1f}h | "
            f"¥{s.regular_cost:,.2f} | ¥{s.overtime_cost:,.2f} | "
            f"¥{s.total_cost:,.2f} | {s.employee_count} |"
        )
    lines.append("")

    lines.append("## 员工成本明细\n")
    lines.append("| 员工 | 门店 | 岗位 | 总工时 | 加班工时 | 总成本 |")
    lines.append("|------|------|------|--------|----------|--------|")
    for s in emp_summaries:
        lines.append(
            f"| {s.employee_name} | {s.store} | {s.position} | "
            f"{s.total_hours:.1f}h | {s.overtime_hours:.1f}h | "
            f"¥{s.total_cost:,.2f} |"
        )
    lines.append("")

    content = "\n".join(lines)
    Path(output_path).write_text(content, encoding="utf-8")
