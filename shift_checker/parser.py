from __future__ import annotations

import datetime
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from shift_checker.config import (
    build_column_lookup,
    build_employee_lookup,
    load_column_synonyms,
    load_employees,
    load_shift_codes,
)

logger = logging.getLogger(__name__)


@dataclass
class ShiftRecord:
    employee_id: str
    employee_name: str
    store: str
    date: datetime.date
    shift_code: str
    start_time: datetime.time | None
    end_time: datetime.time | None
    hours: float
    overnight: bool
    source_file: str
    row_number: int


def _parse_time(val: Any) -> datetime.time | None:
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val
    if isinstance(val, datetime.datetime):
        return val.time()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _parse_date(val: Any) -> datetime.date | None:
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.match(r"(\d{1,2})月(\d{1,2})日", s)
    if m:
        return datetime.date(2025, int(m.group(1)), int(m.group(2)))
    try:
        return datetime.datetime.strptime(s, "%Y%m%d").date()
    except ValueError:
        pass
    return None


def _resolve_shift(
    shift_val: Any,
    start_val: Any,
    end_val: Any,
    shift_codes: dict[str, dict],
) -> tuple[str, datetime.time | None, datetime.time | None, float, bool]:
    shift_str = str(shift_val).strip() if shift_val is not None else ""
    code_info = shift_codes.get(shift_str)

    if code_info and code_info.get("start") and code_info.get("end"):
        start = _parse_time(code_info["start"])
        end = _parse_time(code_info["end"])
        hours = code_info.get("hours", 8.0)
        overnight = code_info.get("overnight", False)
        return shift_str, start, end, hours, overnight

    start = _parse_time(start_val)
    end = _parse_time(end_val)
    if start and end:
        if end > start:
            delta = datetime.datetime.combine(datetime.date.today(), end) - datetime.datetime.combine(datetime.date.today(), start)
            hours = delta.total_seconds() / 3600
            overnight = False
        else:
            delta = datetime.datetime.combine(datetime.date.today(), end) + datetime.timedelta(days=1) - datetime.datetime.combine(datetime.date.today(), start)
            hours = delta.total_seconds() / 3600
            overnight = True
        return shift_str, start, end, hours, overnight

    if code_info:
        start = _parse_time(code_info.get("start"))
        end = _parse_time(code_info.get("end"))
        hours = code_info.get("hours", 0)
        overnight = code_info.get("overnight", False)
        return shift_str, start, end, hours, overnight

    return shift_str, start, end, 0.0, False


def _find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row + 1, 20)):
        vals = [str(ws.cell(row=row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if len(vals) >= 3 and sum(1 for v in vals if v) >= 3:
            return row
    return 1


def _detect_columns(headers: list[str], column_lookup: dict[str, str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, h in enumerate(headers):
        h_clean = str(h).strip()
        if h_clean in column_lookup:
            canonical = column_lookup[h_clean]
            if canonical not in mapping:
                mapping[canonical] = idx
    return mapping


def _expand_merged_rows(ws, header_row: int) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merge_range in ws.merged_cells.ranges:
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        if min_row <= header_row:
            continue
        val = ws.cell(row=min_row, column=min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                merged_values[(r, c)] = val
    return merged_values


def _get_cell_value(ws, row: int, col: int, merged_values: dict) -> Any:
    if (row, col) in merged_values:
        return merged_values[(row, col)]
    cell = ws.cell(row=row, column=col)
    return cell.value


def _extract_store_name(filename: str) -> str:
    stem = Path(filename).stem
    parts = stem.split("_")
    if parts:
        return parts[0]
    return stem


def parse_workbook(
    filepath: str | Path,
    column_lookup: dict[str, str],
    shift_codes: dict[str, dict],
    emp_lookup: dict,
) -> list[ShiftRecord]:
    filepath = Path(filepath)
    store_name = _extract_store_name(filepath.name)
    records: list[ShiftRecord] = []

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        logger.warning("Skipping unreadable workbook %s: %s", filepath, e)
        return records

    if not wb.sheetnames:
        logger.warning("No sheets in %s, skipping", filepath)
        return records

    ws = wb.active
    if ws.max_row is None or ws.max_row < 2:
        logger.warning("Empty or near-empty sheet in %s, skipping", filepath)
        return records

    header_row = _find_header_row(ws)
    raw_headers = [_get_cell_value(ws, header_row, c, {}) for c in range(1, ws.max_column + 1)]

    col_map = _detect_columns(raw_headers, column_lookup)

    date_headers_in_cols: dict[int, datetime.date] = {}
    for col_idx, h in enumerate(raw_headers):
        d = _parse_date(h)
        if d:
            date_headers_in_cols[col_idx] = d

    has_date_column = "date" in col_map
    is_wide_format = bool(date_headers_in_cols) and not has_date_column

    if is_wide_format:
        required_wide = {"employee_id"}
    else:
        required_wide = {"employee_id", "date"}

    if not required_wide.issubset(col_map.keys()):
        logger.warning(
            "Missing required columns %s in %s (found %s), skipping",
            required_wide - col_map.keys(),
            filepath,
            col_map.keys(),
        )
        return records

    merged_values = _expand_merged_rows(ws, header_row)

    if is_wide_format:
        records = _parse_wide_format(
            ws, header_row, col_map, column_lookup, shift_codes, emp_lookup,
            store_name, str(filepath), merged_values, date_headers_in_cols,
        )
    else:
        records = _parse_long_format(
            ws, header_row, col_map, column_lookup, shift_codes, emp_lookup,
            store_name, str(filepath), merged_values,
        )

    wb.close()
    return records


def _parse_long_format(
    ws, header_row, col_map, column_lookup, shift_codes, emp_lookup,
    store_name, source_file, merged_values,
) -> list[ShiftRecord]:
    records: list[ShiftRecord] = []

    emp_id_col = col_map.get("employee_id")
    emp_name_col = col_map.get("employee_name")
    date_col = col_map.get("date")
    shift_col = col_map.get("shift")
    start_col = col_map.get("start_time")
    end_col = col_map.get("end_time")

    for row in range(header_row + 1, ws.max_row + 1):
        raw_id = _get_cell_value(ws, row, emp_id_col + 1, merged_values) if emp_id_col is not None else None
        if raw_id is None:
            continue
        emp_id = str(raw_id).strip()
        if not emp_id:
            continue

        raw_name = _get_cell_value(ws, row, emp_name_col + 1, merged_values) if emp_name_col is not None else ""
        emp_name = str(raw_name).strip() if raw_name else ""

        resolved_id = emp_lookup.get("alias_to_id", {}).get(emp_name, emp_id)
        resolved_name = emp_name
        if resolved_id in emp_lookup.get("by_id", {}):
            resolved_name = emp_lookup["by_id"][resolved_id]["employee_name"]

        raw_date = _get_cell_value(ws, row, date_col + 1, merged_values) if date_col is not None else None
        parsed_date = _parse_date(raw_date)
        if parsed_date is None:
            continue

        raw_shift = _get_cell_value(ws, row, shift_col + 1, merged_values) if shift_col is not None else None
        raw_start = _get_cell_value(ws, row, start_col + 1, merged_values) if start_col is not None else None
        raw_end = _get_cell_value(ws, row, end_col + 1, merged_values) if end_col is not None else None

        shift_code, start_time, end_time, hours, overnight = _resolve_shift(
            raw_shift, raw_start, raw_end, shift_codes,
        )

        if shift_code.upper() in ("OFF", "休") or (not start_time and not end_time and hours == 0):
            continue

        records.append(ShiftRecord(
            employee_id=resolved_id,
            employee_name=resolved_name,
            store=store_name,
            date=parsed_date,
            shift_code=shift_code,
            start_time=start_time,
            end_time=end_time,
            hours=hours,
            overnight=overnight,
            source_file=source_file,
            row_number=row,
        ))

    return records


def _parse_wide_format(
    ws, header_row, col_map, column_lookup, shift_codes, emp_lookup,
    store_name, source_file, merged_values, date_headers_in_cols,
) -> list[ShiftRecord]:
    records: list[ShiftRecord] = []

    emp_id_col = col_map.get("employee_id")
    emp_name_col = col_map.get("employee_name")

    for row in range(header_row + 1, ws.max_row + 1):
        raw_id = _get_cell_value(ws, row, emp_id_col + 1, merged_values) if emp_id_col is not None else None
        if raw_id is None:
            continue
        emp_id = str(raw_id).strip()
        if not emp_id:
            continue

        raw_name = _get_cell_value(ws, row, emp_name_col + 1, merged_values) if emp_name_col is not None else ""
        emp_name = str(raw_name).strip() if raw_name else ""

        resolved_id = emp_lookup.get("alias_to_id", {}).get(emp_name, emp_id)
        resolved_name = emp_name
        if resolved_id in emp_lookup.get("by_id", {}):
            resolved_name = emp_lookup["by_id"][resolved_id]["employee_name"]

        for col_idx, parsed_date in date_headers_in_cols.items():
            raw_val = _get_cell_value(ws, row, col_idx + 1, merged_values)
            if raw_val is None:
                continue

            shift_code, start_time, end_time, hours, overnight = _resolve_shift(
                raw_val, None, None, shift_codes,
            )

            if shift_code.upper() in ("OFF", "休") or (not start_time and not end_time and hours == 0):
                continue

            records.append(ShiftRecord(
                employee_id=resolved_id,
                employee_name=resolved_name,
                store=store_name,
                date=parsed_date,
                shift_code=shift_code,
                start_time=start_time,
                end_time=end_time,
                hours=hours,
                overnight=overnight,
                source_file=source_file,
                row_number=row,
            ))

    return records


def parse_all_workbooks(
    directory: str | Path,
    column_synonyms: dict | None = None,
    shift_codes: dict | None = None,
    employees: list[dict] | None = None,
) -> list[ShiftRecord]:
    directory = Path(directory)
    if column_synonyms is None:
        column_synonyms = load_column_synonyms()
    if shift_codes is None:
        shift_codes = load_shift_codes()
    if employees is None:
        employees = load_employees()

    column_lookup = build_column_lookup(column_synonyms)
    emp_lookup = build_employee_lookup(employees)

    all_records: list[ShiftRecord] = []
    xlsx_files = sorted(directory.glob("*.xlsx"))

    if not xlsx_files:
        logger.warning("No xlsx files found in %s", directory)

    for f in xlsx_files:
        if f.name.startswith("~$"):
            continue
        logger.info("Parsing %s", f.name)
        recs = parse_workbook(f, column_lookup, shift_codes, emp_lookup)
        all_records.extend(recs)
        logger.info("  -> %d records", len(recs))

    return all_records
