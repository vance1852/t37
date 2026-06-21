from __future__ import annotations

import datetime
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from shift_checker.config import (
    build_column_lookup,
    build_employee_lookup,
    load_column_synonyms,
    load_employees,
    load_holidays,
    load_rules,
    load_shift_codes,
    load_cost_rates,
)


@pytest.fixture
def column_synonyms():
    return load_column_synonyms()


@pytest.fixture
def column_lookup(column_synonyms):
    return build_column_lookup(column_synonyms)


@pytest.fixture
def shift_codes():
    return load_shift_codes()


@pytest.fixture
def employees():
    return load_employees()


@pytest.fixture
def emp_lookup(employees):
    return build_employee_lookup(employees)


@pytest.fixture
def holidays():
    return load_holidays()


@pytest.fixture
def rules():
    return load_rules()


@pytest.fixture
def cost_rates():
    return load_cost_rates()


def _make_wb(headers: list[str], data: list[list], merged_ranges: list[str] | None = None):
    wb = Workbook()
    ws = wb.active
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    if merged_ranges:
        for mr in merged_ranges:
            ws.merge_cells(mr)
    return wb


def _save_wb(wb, tmp_path: Path, name: str = "test.xlsx") -> Path:
    p = tmp_path / name
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def wb_long_format():
    headers = ["员工号", "姓名", "日期", "班次", "上班时间", "下班时间"]
    data = [
        ["E001", "张三", datetime.date(2025, 1, 6), "A", "09:00", "17:00"],
        ["E001", "张三", datetime.date(2025, 1, 7), "A", "09:00", "17:00"],
        ["E002", "李四", datetime.date(2025, 1, 6), "B", "13:00", "21:00"],
    ]
    return _make_wb(headers, data)


@pytest.fixture
def wb_shift_codes_only():
    headers = ["工号", "名字", "工作日", "班别"]
    data = [
        ["E001", "张三", "2025/1/6", "早"],
        ["E001", "张三", "2025/1/7", "晚"],
        ["E001", "张三", "2025/1/8", "休"],
    ]
    return _make_wb(headers, data)


@pytest.fixture
def wb_overnight():
    headers = ["员工编号", "姓名", "日期", "班次", "上班时间", "下班时间"]
    data = [
        ["E001", "张三", datetime.date(2025, 1, 6), "C", "21:00", "05:00"],
        ["E001", "张三", datetime.date(2025, 1, 7), "A", "09:00", "17:00"],
    ]
    return _make_wb(headers, data)


@pytest.fixture
def wb_merged_cells():
    headers = ["员工号", "姓名", "1月6日", "1月7日"]
    data = [
        ["E001", "张三", "早", "中"],
        ["E002", "李四", "中", "早"],
    ]
    merged = ["A2:A3"]
    return _make_wb(headers, data, merged)


@pytest.fixture
def wb_alias_trap():
    headers = ["EmpID", "Name", "Date", "Shift"]
    data = [
        ["E001", "zhangsan", datetime.date(2025, 1, 6), "A"],
        ["E001", "张三", datetime.date(2025, 1, 7), "A"],
    ]
    return _make_wb(headers, data)


@pytest.fixture
def wb_rest_insufficient():
    headers = ["员工号", "姓名", "日期", "班次", "上班时间", "下班时间"]
    data = [
        ["E001", "张三", datetime.date(2025, 1, 6), "", "21:00", "05:00"],
        ["E001", "张三", datetime.date(2025, 1, 7), "", "08:00", "16:00"],
    ]
    return _make_wb(headers, data)


@pytest.fixture
def wb_holiday():
    headers = ["员工号", "姓名", "日期", "班次", "上班时间", "下班时间"]
    data = [
        ["E001", "张三", datetime.date(2025, 1, 1), "A", "09:00", "17:00"],
    ]
    return _make_wb(headers, data)


@pytest.fixture
def wb_corrupt(tmp_path):
    p = tmp_path / "corrupt.xlsx"
    p.write_bytes(b"not a real excel file")
    return p


@pytest.fixture
def wb_no_required_cols():
    headers = ["门店", "备注"]
    data = [["旗舰店", "测试"]]
    return _make_wb(headers, data)


@pytest.fixture
def tmp_output_dir(tmp_path):
    d = tmp_path / "output"
    d.mkdir()
    return d
