from __future__ import annotations

import datetime
import tempfile
from pathlib import Path

import pytest

from shift_checker.parser import (
    ShiftRecord,
    _parse_date,
    _parse_time,
    _resolve_shift,
    parse_all_workbooks,
    parse_workbook,
)


class TestParseTime:
    def test_time_object(self):
        t = datetime.time(9, 30)
        assert _parse_time(t) == t

    def test_datetime_object(self):
        dt = datetime.datetime(2025, 1, 1, 14, 0)
        assert _parse_time(dt) == datetime.time(14, 0)

    def test_string_hh_mm(self):
        assert _parse_time("09:30") == datetime.time(9, 30)

    def test_string_hh_mm_ss(self):
        assert _parse_time("14:30:00") == datetime.time(14, 30)

    def test_none(self):
        assert _parse_time(None) is None

    def test_empty_string(self):
        assert _parse_time("") is None

    def test_invalid_string(self):
        assert _parse_time("abc") is None


class TestParseDate:
    def test_date_object(self):
        d = datetime.date(2025, 1, 15)
        assert _parse_date(d) == d

    def test_datetime_object(self):
        dt = datetime.datetime(2025, 1, 15, 10, 0)
        assert _parse_date(dt) == datetime.date(2025, 1, 15)

    def test_iso_format(self):
        assert _parse_date("2025-01-15") == datetime.date(2025, 1, 15)

    def test_slash_format(self):
        assert _parse_date("2025/1/6") == datetime.date(2025, 1, 6)

    def test_chinese_format(self):
        assert _parse_date("1月15日") == datetime.date(2025, 1, 15)

    def test_none(self):
        assert _parse_date(None) is None

    def test_empty_string(self):
        assert _parse_date("") is None


class TestResolveShift:
    @pytest.fixture
    def codes(self):
        return {
            "A": {"start": "09:00", "end": "17:00", "label": "早班", "hours": 8.0},
            "C": {"start": "21:00", "end": "05:00", "label": "晚班", "hours": 8.0, "overnight": True},
            "OFF": {"start": None, "end": None, "label": "休息", "hours": 0},
        }

    def test_shift_code_resolves(self, codes):
        code, start, end, hours, overnight = _resolve_shift("A", None, None, codes)
        assert code == "A"
        assert start == datetime.time(9, 0)
        assert end == datetime.time(17, 0)
        assert hours == 8.0
        assert overnight is False

    def test_overnight_code(self, codes):
        code, start, end, hours, overnight = _resolve_shift("C", None, None, codes)
        assert overnight is True
        assert hours == 8.0

    def test_explicit_times_override(self, codes):
        code, start, end, hours, overnight = _resolve_shift("", "22:00", "06:00", codes)
        assert overnight is True
        assert hours == 8.0

    def test_off_shift(self, codes):
        code, start, end, hours, overnight = _resolve_shift("OFF", None, None, codes)
        assert hours == 0


class TestColumnRecognition:
    def test_recognizes_synonyms(self, column_lookup):
        assert column_lookup["员工号"] == "employee_id"
        assert column_lookup["工号"] == "employee_id"
        assert column_lookup["EmpID"] == "employee_id"
        assert column_lookup["姓名"] == "employee_name"
        assert column_lookup["名字"] == "employee_name"
        assert column_lookup["日期"] == "date"
        assert column_lookup["工作日"] == "date"
        assert column_lookup["班次"] == "shift"
        assert column_lookup["班别"] == "shift"

    def test_unknown_column_not_in_lookup(self, column_lookup):
        assert "未知列名" not in column_lookup


class TestParseWorkbook:
    def test_long_format(self, wb_long_format, tmp_path, column_lookup, shift_codes, emp_lookup):
        p = tmp_path / "long.xlsx"
        wb_long_format.save(p)
        wb_long_format.close()
        recs = parse_workbook(p, column_lookup, shift_codes, emp_lookup)
        assert len(recs) == 3
        assert recs[0].employee_id == "E001"
        assert recs[0].employee_name == "张三"
        assert recs[0].hours == 8.0

    def test_shift_codes_only(self, wb_shift_codes_only, tmp_path, column_lookup, shift_codes, emp_lookup):
        p = tmp_path / "codes.xlsx"
        wb_shift_codes_only.save(p)
        wb_shift_codes_only.close()
        recs = parse_workbook(p, column_lookup, shift_codes, emp_lookup)
        assert len(recs) == 2
        assert all(r.employee_id == "E001" for r in recs)
        overnight_recs = [r for r in recs if r.overnight]
        assert len(overnight_recs) == 1

    def test_overnight_shift(self, wb_overnight, tmp_path, column_lookup, shift_codes, emp_lookup):
        p = tmp_path / "overnight.xlsx"
        wb_overnight.save(p)
        wb_overnight.close()
        recs = parse_workbook(p, column_lookup, shift_codes, emp_lookup)
        assert len(recs) == 2
        night = recs[0]
        assert night.overnight is True
        assert night.start_time == datetime.time(21, 0)
        assert night.end_time == datetime.time(5, 0)

    def test_merged_cells(self, wb_merged_cells, tmp_path, column_lookup, shift_codes, emp_lookup):
        p = tmp_path / "merged.xlsx"
        wb_merged_cells.save(p)
        wb_merged_cells.close()
        recs = parse_workbook(p, column_lookup, shift_codes, emp_lookup)
        assert len(recs) >= 3
        e001_recs = [r for r in recs if r.employee_id == "E001"]
        assert len(e001_recs) >= 2

    def test_alias_resolution(self, wb_alias_trap, tmp_path, column_lookup, shift_codes, emp_lookup):
        p = tmp_path / "alias.xlsx"
        wb_alias_trap.save(p)
        wb_alias_trap.close()
        recs = parse_workbook(p, column_lookup, shift_codes, emp_lookup)
        assert len(recs) == 2
        for r in recs:
            assert r.employee_id == "E001"
            assert r.employee_name == "张三"

    def test_corrupt_workbook_skipped(self, wb_corrupt, column_lookup, shift_codes, emp_lookup):
        recs = parse_workbook(wb_corrupt, column_lookup, shift_codes, emp_lookup)
        assert recs == []

    def test_no_required_columns_skipped(self, wb_no_required_cols, tmp_path, column_lookup, shift_codes, emp_lookup):
        p = tmp_path / "nocols.xlsx"
        wb_no_required_cols.save(p)
        wb_no_required_cols.close()
        recs = parse_workbook(p, column_lookup, shift_codes, emp_lookup)
        assert recs == []


class TestParseAllWorkbooks:
    def test_parses_directory(self, tmp_path, column_lookup, shift_codes, emp_lookup):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="员工号")
        ws.cell(row=1, column=2, value="姓名")
        ws.cell(row=1, column=3, value="日期")
        ws.cell(row=1, column=4, value="班次")
        ws.cell(row=2, column=1, value="E001")
        ws.cell(row=2, column=2, value="张三")
        ws.cell(row=2, column=3, value=datetime.date(2025, 1, 6))
        ws.cell(row=2, column=4, value="A")
        p = tmp_path / "store.xlsx"
        wb.save(p)
        wb.close()

        recs = parse_all_workbooks(tmp_path)
        assert len(recs) == 1
